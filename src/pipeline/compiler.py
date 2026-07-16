"""
Stage 3: Compile all ModuleOutputs into a SurveyCTO-ready .xlsx file.

Sheets produced:
    survey   — fixed_start rows, module rows, fixed_end rows
    choices  — fixed_choices rows, then all module choice rows
    settings — header row only, values left blank for manual entry
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models.survey import ChoiceRow, ModuleOutput, SurveyRow
from utils.logger import get_logger
from config.compiler import (SETTINGS_HEADERS, SURVEY_WIDTHS, SURVEY_PREFIX_WIDTHS, 
                             CHOICES_WIDTHS, CHOICES_PREFIX_WIDTHS, FIXED_ROWS_PATH
                             ,_WRAP, _HEADER_FONT, _CELL_FONT)

log = get_logger(__name__)


def _col_width(header: str, fixed: dict[str, float], prefix: dict[str, float]) -> float:
    if header in fixed:
        return fixed[header]
    for pfx, w in prefix.items():
        if header.startswith(pfx):
            return w
    return 36


# ---------------------------------------------------------------------------
# Fixed rows loader
# ---------------------------------------------------------------------------

def _load_fixed_rows(languages: list[str], path: Path = FIXED_ROWS_PATH) -> dict:
    """
    Load fixed_rows.yaml and materialise SurveyRow / ChoiceRow objects,
    mapping stored language keys to the form's language list.
    Labels for languages not present in the YAML are set to "".
    """
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))

    def _labels(stored: dict[str, str] | None, prefix: str) -> dict[str, str]:
        stored = stored or {}
        return {f"{prefix}{lang}": stored.get(lang, "") for lang in languages}

    def _to_survey_row(d: dict[str, Any]) -> SurveyRow:
        return SurveyRow(
            type=d.get("type", ""),
            name=d.get("name", ""),
            labels=_labels(d.get("labels"), "label:"),
            appearance=d.get("appearance", ""),
            relevance=d.get("relevance", ""),
            required=d.get("required", ""),
            hints=_labels(d.get("hints"), "hint:"),
            constraint=d.get("constraint", ""),
            constraint_messages=_labels(d.get("constraint_messages"), "constraint message:"),
            calculation=d.get("calculation", ""),
            choice_filter=d.get("choice_filter", ""),
            read_only=d.get("read_only", ""),
            default=d.get("default", ""),
            repeat_count=d.get("repeat_count", ""),
            media_images=_labels(d.get("media_images"), "media:image:"),
            minimum_seconds=d.get("minimum_seconds", ""),
            publishable=d.get("publishable", ""),
        )

    def _to_choice_row(d: dict[str, Any]) -> ChoiceRow:
        return ChoiceRow(
            list_name=str(d.get("list_name", "")),
            value=int(d["value"]),
            labels=_labels(d.get("labels"), "label:"),
            filter=d.get("filter"),
        )

    return {
        "survey_start": [_to_survey_row(r) for r in raw.get("survey_start", [])],
        "survey_end":   [_to_survey_row(r) for r in raw.get("survey_end", [])],
        "fixed_choices":[_to_choice_row(r) for r in raw.get("fixed_choices", [])],
    }


# ---------------------------------------------------------------------------
# Column resolution
# ---------------------------------------------------------------------------

def _resolve_survey_headers(
    languages: list[str],
) -> list[str]:
    # Derive language-keyed columns from languages list directly (don't rely on
    # first row having all keys populated)
    label_cols              = [f"label:{l.lower()}" for l in languages]
    hint_cols               = [f"hint:{l.lower()}" for l in languages]
    constraint_message_cols = [f"constraint message:{l.lower()}" for l in languages]
    media_image_cols        = [f"media:image:{l.lower()}" for l in languages]

    return (
        ["type", "name"]
        + label_cols
        + ["appearance", "relevance", "required"]
        + hint_cols
        + ["constraint"]
        + constraint_message_cols
        + ["calculation", "choice_filter", "read only", "default", "repeat_count"]
        + media_image_cols
        + ["minimum_seconds", "publishable"]
    )


def _resolve_choices_headers(languages: list[str]) -> list[str]:
    return ["list_name", "value"] + [f"label:{l}" for l in languages] + ["filter"]


# ---------------------------------------------------------------------------
# Row serialization
# ---------------------------------------------------------------------------

def _survey_row_to_list(row: SurveyRow, headers: list[str]) -> list[str]:
    mapping: dict[str, str] = {
        "type": row.type,
        "name": row.name,
        "appearance": row.appearance,
        "relevance": row.relevance,
        "required": row.required,
        "constraint": row.constraint,
        "calculation": row.calculation,
        "choice_filter": row.choice_filter,
        "read only": row.read_only,
        "default": row.default,
        "repeat_count": row.repeat_count,
        "minimum_seconds": row.minimum_seconds,
        "publishable": row.publishable,
        **row.labels,
        **row.hints,
        **row.constraint_messages,
        **row.media_images,
    }
    return [mapping.get(h, "") for h in headers]


def _choice_row_to_list(row: ChoiceRow, headers: list[str]) -> list[str]:
    mapping: dict = {
        "list_name": row.list_name,
        "value": row.value,
        "filter": row.filter if row.filter is not None else "",
        **row.labels,
    }
    return [mapping.get(h, "") for h in headers]


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------



def _apply_sheet_formatting(
    ws,
    headers: list[str],
    fixed_widths: dict[str, float],
    prefix_widths: dict[str, float],
) -> None:
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _col_width(
            header, fixed_widths, prefix_widths
        )
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = _WRAP
            cell.font = _HEADER_FONT if cell.row == 1 else _CELL_FONT


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_survey_sheet(
    ws,
    outputs: list[ModuleOutput],
    headers: list[str],
    survey_start: list[SurveyRow],
    survey_end: list[SurveyRow],
) -> None:
    ws.append(headers)

    for row in survey_start:
        ws.append(_survey_row_to_list(row, headers))

    for output in outputs:
        for row in output.survey_rows:
            ws.append(_survey_row_to_list(row, headers))

    for row in survey_end:
        ws.append(_survey_row_to_list(row, headers))

    _apply_sheet_formatting(ws, headers, SURVEY_WIDTHS, SURVEY_PREFIX_WIDTHS)


def _deduplicate_choices(rows: list[ChoiceRow]) -> list[ChoiceRow]:
    seen: set[tuple[str, int]] = set()
    result: list[ChoiceRow] = []
    for row in rows:
        key = (row.list_name, row.value)
        if key not in seen:
            seen.add(key)
            result.append(row)
    return result


def _write_choices_sheet(
    ws,
    outputs: list[ModuleOutput],
    headers: list[str],
    fixed_choices: list[ChoiceRow],
) -> None:
    ws.append(headers)

    # Fixed choices first, then module choices — deduplicate across both
    all_rows = fixed_choices + [r for o in outputs for r in o.choice_rows]
    deduped  = _deduplicate_choices(all_rows)

    current_list: str | None = None
    for row in deduped:
        if row.list_name != current_list:
            if current_list is not None:
                ws.append([""] * len(headers))
            current_list = row.list_name
        ws.append(_choice_row_to_list(row, headers))
    if current_list is not None:
        ws.append([""] * len(headers))

    _apply_sheet_formatting(ws, headers, CHOICES_WIDTHS, CHOICES_PREFIX_WIDTHS)


def _write_settings_sheet(ws) -> None:
    ws.append(SETTINGS_HEADERS)
    ws.append([""] * len(SETTINGS_HEADERS))
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.alignment = _WRAP


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def compile_outputs(
    outputs: list[ModuleOutput],
    docx_path: Path,
    languages: list[str],
    out_dir: Path | None = None,
    fixed_rows_path: Path = FIXED_ROWS_PATH,
) -> Path:
    """
    Compile all ModuleOutputs into a SurveyCTO .xlsx file.

    Args:
        outputs:          List of ModuleOutput from parser.parse_all_modules().
        docx_path:        Original .docx path — used to derive output filename.
        languages:        Ordered list of language keys, e.g. ["english", "urdu"].
        out_dir:          Output directory. Defaults to docx_path's parent.
        fixed_rows_path:  Override path to fixed_rows.yaml.

    Returns:
        Path to the written .xlsx file.
    """
    out_dir  = out_dir or docx_path.parent
    out_path = out_dir / (docx_path.stem + ".xlsx")

    fixed = _load_fixed_rows(languages, fixed_rows_path)

    survey_headers  = _resolve_survey_headers(languages)
    choices_headers = _resolve_choices_headers(languages)

    wb = Workbook()

    ws_survey = wb.create_sheet("survey")
    _write_survey_sheet(
        ws_survey, outputs, survey_headers,
        fixed["survey_start"], fixed["survey_end"],
    )

    ws_choices = wb.create_sheet("choices")
    _write_choices_sheet(ws_choices, outputs, choices_headers, fixed["fixed_choices"])

    ws_settings = wb.create_sheet("settings")
    _write_settings_sheet(ws_settings)

    del wb[wb.sheetnames[0]]

    wb.save(out_path)
    log.info(f"Compiled {len(outputs)} modules → {out_path}")
    return out_path